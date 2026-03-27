#!/usr/bin/env python3

import argparse, json, requests, getpass, re, datetime
from openpyxl import load_workbook, Workbook
from cvss import CVSS3

# Disable selfsigned certs triggering alerts
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

USER_AGENT="pwn2ex v0.1"

class PwnResponse():
    def __init__(self, status: str, datas):
        self.status = status
        self.datas = datas

class Vulnerability():
    def __init__(self, id: int, description: str, criticality: str, cvss: float, assets: list[str], detection_date: str, root_cause: str, corrective_action: str, close_date: str, evidence: str, active: str, observation: str):
        self.id = id
        self.description = description
        self.criticality = criticality
        self.cvss = cvss
        self.assets = assets
        self.detection_date = detection_date
        self.root_cause = root_cause
        self.corrective_action = corrective_action
        self.close_date = close_date
        self.evidence = evidence
        self.active = active
        self.observation = observation

class Audit():
    def __init__(self, id: str, title: str, auditType: str, findings: list[Vulnerability], company: str):
        self.id = id
        self.title = title
        self.auditType = auditType
        self.findings = findings
        self.company = company
    def __str__(self):
        return f"{self.title} ({self.company})"


def main():
    p = argparse.ArgumentParser(description="automatically write vulnerabilities from a pwndoc audit to a template excel file")
    p.add_argument("-i","--input", required=True, help="input excel template")
    p.add_argument("-o","--output", required=True, help="output processed excel file")
    p.add_argument("-r","--row", required=False, help="specify starting row (default: 4)", type=int, default=4)
    p.add_argument("-c","--column", required=False, help="specify starting column (default: B)", type=str, default="B")
    p.add_argument("--keep-metadata", required=False, help="keep output excel file metadata (default: False)", action="store_true", default=False)
    p.add_argument("target", help="target pwndoc server")
    args = p.parse_args()

    username = input("Username: ")
    password = getpass.getpass(prompt="Password: ")
    totp = input("TOTP Token (Leave empty if none): ")
    target = args.target.rstrip('/')

    login = auth(target, username, password, totp).datas
    token = 'JWT%20' + login["token"]
    audit = get_audit(target=target, token=token)
    save_audit(audit=audit, template=args.input, output=args.output, row=args.row, col=args.column.upper(), wipe=not args.keep_metadata)


def get_audit(target: str, token: str):
    url_audits = target + "/api/audits"
    audits_resp = req(url=url_audits, token=token).datas
    choosable_audits = []
    for a in audits_resp:
        choosable_audits.append(
            Audit(
                id=a["_id"],
                title=a["name"],
                auditType=a["auditType"],
                company=a["company"]["name"],
                findings=[]
            )
        )
    log("Choose the audit you wish to export:")
    chosen_audit = list_choice(choosable_audits)
    url_audit = target + "/api/audits/" + chosen_audit.id
    audit_resp = req(url=url_audit, token=token).datas

    vulns = []
    for vuln in audit_resp["findings"]:
        cvss = vuln.get("cvssv3")
        score = -1
        criticality = "None"
        if cvss is not None:
            vector = CVSS3(cvss)
            score = vector.scores()[0]
            criticality = vector.severities()[0]
        vulns.append(
            Vulnerability(
                id=vuln.get("identifier"),
                description=vuln.get("title"),
                criticality=criticality,
                cvss=score if score != -1 else 0,
                root_cause="N/A",
                assets=strip_html_to_list(vuln.get("scope"))[:-1],
                detection_date=audit_resp.get("date_end"),
                observation=strip_html(vuln.get("observation")),
                corrective_action=strip_html(vuln.get("remediation")),
                close_date="TBD",
                evidence=f"Graphic evidences: {audit_resp.get("name")}",
                active="YES"
            )
        )
    final_audit = Audit(
        id=audit_resp["_id"],
        title=audit_resp["name"],
        auditType=audit_resp["auditType"],
        findings=vulns,
        company=audit_resp["company"]["name"]
    )
    return final_audit


def strip_html_to_list(s):
    if s is None:
        s = ""
    semistripped = re.sub(r"</.*?>", " ", s)
    stripped = re.sub(r"<.*?>", "", semistripped).split(" ")
    return stripped


def strip_html(s):
    if s is None:
        s = ""
    return re.sub(r"<.*?>", "", s)


def save_audit(audit: Audit, template: str, output: str, row: int, col: str, wipe: bool):
    wb = load_workbook(template)
    log("Select the sheet to fill:")
    sheet = wb[list_choice(wb.sheetnames)]
    for i, vuln in enumerate(audit.findings):
        sheet[f"{col}{row}"].value = i+1
        sheet[f"{next_col(col, 1)}{row}"].value = vuln.description
        sheet[f"{next_col(col, 2)}{row}"].value = vuln.criticality
        sheet[f"{next_col(col, 3)}{row}"].value = vuln.cvss
        sheet[f"{next_col(col, 4)}{row}"].value = "\n".join(vuln.assets)
        sheet[f"{next_col(col, 5)}{row}"].value = vuln.detection_date
        sheet[f"{next_col(col, 6)}{row}"].value = vuln.root_cause
        sheet[f"{next_col(col, 7)}{row}"].value = vuln.corrective_action
        sheet[f"{next_col(col, 8)}{row}"].value = vuln.close_date
        sheet[f"{next_col(col, 9)}{row}"].value = vuln.evidence
        sheet[f"{next_col(col, 10)}{row}"].value = vuln.active
        sheet[f"{next_col(col, 11)}{row}"].value = vuln.observation
        row+=1
    if wipe:
        wipe_metadata(wb)
    wb.save(output)
    log(f"File \"{output}\" created successfully!")


def wipe_metadata(wb: Workbook):
    p = wb.properties
    p.creator=None
    p.title=None
    p.description=None
    p.subject=None
    p.identifier=None
    p.language=None
    p.created=datetime.datetime(1970, 1, 1, 0, 0, 0)
    p.modified=datetime.datetime(1970, 1, 1, 0, 0, 0)
    p.lastModifiedBy=None
    p.category=None
    p.contentStatus=None
    p.version=None
    p.revision=None
    p.keywords=None
    p.lastPrinted=None


def next_col(letter: str, jump: int):
    next = chr(ord(letter)+jump)
    return next


def list_choice(array: list):
    for i, v in enumerate(array):
        print(f"[{i}] {v}")
    try:
        id_item = int(input("Choice (ID): "))
        item = array[id_item]
    except Exception as e:
        raise e
    return item


def auth(target: str, username: str, password: str, totp: str):
    url = target + "/api/users/token"
    creds = {
        "username": username,
        "password": password
    }
    if totp is not None:
        creds["totpToken"] = totp

    resp = requests.post(url, headers={
        "accept": "application/json",
        "Content-Type": "application/json",
        "User-Agent": USER_AGENT
    },
    json=creds, verify=False)
    contents = json.loads(resp.content)
    status = contents.get("status")
    datas = contents.get("datas")
    return check_success(PwnResponse(status=status, datas=datas))


def req(url: str, token: str):
    resp = requests.get(url, headers={
        "accept": "application/json",
        "Content-Type": "application/json",
        "User-Agent": USER_AGENT
    },
    cookies={'token': token}, verify=False)
    contents = json.loads(resp.content)
    status = contents.get("status")
    datas = contents.get("datas")
    return check_success(PwnResponse(status=status, datas=datas))


def check_success(response: PwnResponse):
    if response.status != "success":
        err(response.datas)
    return response


def log(msg: str):
    print("\n--------------------")
    print(f"[+] {msg}")
    print("--------------------\n")


def err(msg: str):
    print(f"[-] ERROR: {msg}")
    exit(2)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        log("Bye!")
